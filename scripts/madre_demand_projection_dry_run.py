#!/usr/bin/env python3
"""Dry-run demand projection for Casa de mi Madre stock planning.

The script is read-only against the webapp API. It resolves the workbook
projection against live company IDs, technical sheets, products and inventory
balances, then writes an audit report with suggested internal supply and
purchase residuals.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook


DEFAULT_API_BASE = "https://gestor-estoque-zqw9.onrender.com/api"
DEFAULT_WORKBOOK = "/home/leomassoni/Downloads/Projecao_Demanda_Madre_Abertura.xlsx"
DEFAULT_OUTPUT_DIR = "auditorias"
DEFAULT_COMPANY_ID = 13
DEFAULT_CENTER_NAME = "BAR DE BAIXO"

PROMOTIONAL_AUTHORIAL_UNITS = 140
SAFETY_MARGIN_PERCENT = 20
DEFAULT_FAKE_DEMAND_FACTOR = 0.8

CLOSED_BEVERAGE_ALIASES = {
    "COCA": "COCA COLA LATA",
    "COCA COLA": "COCA COLA LATA",
    "COCA ZERO": "COCA COLA ZERO LATA",
    "COCA COLA ZERO": "COCA COLA ZERO LATA",
    "GUARANA": "GUARANA ANTARCTICA LATA",
    "GUARANA ZERO": "GUARANA ANTARCTICA ZERO LATA",
    "CORONA": "CORONA EXTRA LONG NECK",
    "AGUA PRATA": "AGUA PRATA SEM GAS RETORNAVEL",
    "AGUA PRATA SEM GAS": "AGUA PRATA SEM GAS RETORNAVEL",
    "AGUA PRATA C GAS": "AGUA PRATA COM GAS RETORNAVEL",
    "AGUA PRATA C/ GAS": "AGUA PRATA COM GAS RETORNAVEL",
    "AGUA PRATA COM GAS": "AGUA PRATA COM GAS RETORNAVEL",
}

UNIT_LABELS = {
    "MILLILITER": "ML",
    "GRAM": "G",
    "UNIT": "UN",
    "COMBO": "COMBO",
}

OPERATIONAL_LOCATIONS = {
    "ENTRADA DE PRODUCAO",
    "SAIDA PARA PRODUCAO",
    "SAIDA PARA REQUISICAO",
    "SAIDA POR VENDAS IMPORTADAS",
    "ESTORNO DE VENDAS IMPORTADAS",
    "RECEBIMENTO DE REQUISICAO",
}

OUTBOUND_LOCATIONS = {
    "SAIDA PARA PRODUCAO",
    "SAIDA PARA REQUISICAO",
    "SAIDA POR VENDAS IMPORTADAS",
}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )
    text = re.sub(r"[^A-Za-z0-9 ./_-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.upper()


def parse_decimal(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    compact = re.sub(r"\s+", "", str(value))
    if not compact:
        return None
    normalized = compact
    if "," in compact and "." in compact:
        normalized = compact.replace(".", "").replace(",", ".")
    elif "," in compact:
        normalized = compact.replace(",", ".")
    elif re.match(r"^\d{1,3}(\.\d{3})+$", compact):
        normalized = compact.replace(".", "")
    try:
        parsed = float(normalized)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def format_decimal(value: float, precision: int = 4) -> str:
    rounded = round(value, precision)
    if abs(rounded) < 10 ** -precision:
        rounded = 0
    text = f"{rounded:.{precision}f}".rstrip("0").rstrip(".")
    return text or "0"


def unit_label(unit: str) -> str:
    return UNIT_LABELS.get(unit, unit or "")


def api_get(api_base: str, path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    url = api_base.rstrip("/") + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=90) as response:
        return json.load(response)


def first_numeric_id(values: Iterable[Dict[str, Any]], fallback: int = 0) -> int:
    ids = [int(item["id"]) for item in values if isinstance(item.get("id"), int)]
    return max(ids, default=fallback)


def find_header_row(sheet: Any, required_headers: Sequence[str]) -> int:
    required = {normalize_text(header) for header in required_headers}
    for row_index in range(1, min(sheet.max_row or 1, 20) + 1):
        headers = {normalize_text(cell.value) for cell in sheet[row_index]}
        if required.issubset(headers):
            return row_index
    raise RuntimeError(f"Cabecalho nao encontrado na aba {sheet.title}: {required_headers}")


def iter_table_rows(sheet: Any, required_headers: Sequence[str]) -> Iterable[Dict[str, Any]]:
    header_row = find_header_row(sheet, required_headers)
    headers = [normalize_text(cell.value) for cell in sheet[header_row]]
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in row):
            continue
        yield {headers[index]: value for index, value in enumerate(row) if index < len(headers)}


def read_projection(workbook_path: str) -> Dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    crowd_rows = []
    for row in iter_table_rows(workbook["Projeção de Público"], ["PERIODO", "DIA", "MINIMO", "MAXIMO"]):
        minimum = parse_decimal(row.get("MINIMO")) or 0
        maximum = parse_decimal(row.get("MAXIMO")) or 0
        day = str(row.get("DIA") or "").strip()
        if (not day or normalize_text(day).split(" ")[0] not in {"SEGUNDA", "TERCA", "QUARTA", "QUINTA", "SEXTA", "SABADO", "DOMINGO"}) and minimum <= 0 and maximum <= 0:
            continue
        crowd_rows.append(
            {
                "period": row.get("PERIODO") or "",
                "day": day,
                "minimum": minimum,
                "maximum": maximum,
                "note": row.get("OBSERVACAO") or "",
            }
        )

    authorials = []
    for row in iter_table_rows(workbook["Autorais"], ["DRINK", "PRODUCAO INICIAL", "RESERVA / REPOSICAO", "CAPACIDADE TOTAL"]):
        drink = str(row.get("DRINK") or "").strip()
        if not drink or normalize_text(drink) in {"TOTAL", "NOTAS", "PRODUCAO INICIAL", "RESERVA / REPOSICAO"}:
            continue
        capacity_total = parse_decimal(row.get("CAPACIDADE TOTAL")) or 0
        if capacity_total <= 0:
            continue
        authorials.append(
            {
                "drink": drink,
                "initial": parse_decimal(row.get("PRODUCAO INICIAL")) or 0,
                "reserve": parse_decimal(row.get("RESERVA / REPOSICAO")) or 0,
                "capacity_total": capacity_total,
                "priority": str(row.get("PRIORIDADE") or "").strip(),
            }
        )

    classics = []
    for row in iter_table_rows(workbook["Clássicos"], ["DRINK", "CAPACIDADE"]):
        drink = str(row.get("DRINK") or "").strip()
        if not drink or normalize_text(drink).startswith("TOTAL"):
            continue
        capacity = parse_decimal(row.get("CAPACIDADE")) or 0
        if capacity <= 0:
            continue
        classics.append(
            {
                "drink": drink,
                "capacity": capacity,
                "guideline": row.get("DIRETRIZ") or "",
            }
        )

    closed_beverages = []
    for row in iter_table_rows(workbook["Bebidas Fechadas"], ["PRODUTO", "ESTOQUE RECOMENDADO"]):
        product = str(row.get("PRODUTO") or "").strip()
        if not product:
            continue
        closed_beverages.append(
            {
                "product": product,
                "target_units": parse_decimal(row.get("ESTOQUE RECOMENDADO")) or 0,
                "observed_consumption": parse_decimal(row.get("CONSUMO OBSERVADO")) or 0,
                "note": row.get("OBSERVACAO") or "",
            }
        )

    return {
        "crowd": crowd_rows,
        "authorials": authorials,
        "classics": classics,
        "closed_beverages": closed_beverages,
    }


def next_weekday_on_or_after(start: dt.date, weekday: int) -> dt.date:
    return start + dt.timedelta(days=(weekday - start.weekday()) % 7)


def projection_dates(crowd_rows: Sequence[Dict[str, Any]], first_date: Optional[str]) -> List[Dict[str, Any]]:
    if first_date:
        current = dt.date.fromisoformat(first_date)
    else:
        current = next_weekday_on_or_after(dt.date.today(), 3)

    weekday_by_name = {
        "SEGUNDA": 0,
        "TERCA": 1,
        "QUARTA": 2,
        "QUINTA": 3,
        "SEXTA": 4,
        "SABADO": 5,
        "DOMINGO": 6,
    }
    dated_rows = []
    for row in crowd_rows:
        day_label = normalize_text(row["day"]).split(" ")[0]
        target_weekday = weekday_by_name.get(day_label)
        if target_weekday is not None:
            current = next_weekday_on_or_after(current, target_weekday)
        midpoint = ((row["minimum"] or 0) + (row["maximum"] or 0)) / 2
        dated_rows.append({**row, "date": current.isoformat(), "weight": midpoint})
        current += dt.timedelta(days=1)
    return dated_rows


def fixed_sales_dates(start_date: str, end_date: str) -> List[Dict[str, Any]]:
    start = dt.date.fromisoformat(start_date)
    end = dt.date.fromisoformat(end_date)
    if end < start:
        raise RuntimeError("A data final do relatorio fake nao pode ser anterior a data inicial.")
    rows = []
    current = start
    while current <= end:
        rows.append(
            {
                "period": "Relatorio fake",
                "day": current.strftime("%A"),
                "minimum": 1,
                "maximum": 1,
                "note": "Distribuicao uniforme do relatorio fake",
                "date": current.isoformat(),
                "weight": 1,
            }
        )
        current += dt.timedelta(days=1)
    return rows


def distribute_quantity(total_quantity: float, dated_crowd_rows: Sequence[Dict[str, Any]]) -> List[Tuple[str, float]]:
    total_weight = sum(max(row["weight"], 0) for row in dated_crowd_rows)
    if total_quantity <= 0 or total_weight <= 0:
        return []
    raw = [(row["date"], total_quantity * max(row["weight"], 0) / total_weight) for row in dated_crowd_rows]
    rounded = [(date_key, math.floor(quantity)) for date_key, quantity in raw]
    remainder = int(round(total_quantity - sum(quantity for _, quantity in rounded)))
    fractions = sorted(
        ((index, raw[index][1] - rounded[index][1]) for index in range(len(raw))),
        key=lambda item: item[1],
        reverse=True,
    )
    for index, _fraction in fractions[: max(0, remainder)]:
        rounded[index] = (rounded[index][0], rounded[index][1] + 1)
    return [(date_key, quantity) for date_key, quantity in rounded if quantity > 0]


def product_visible_for_company(product: Dict[str, Any], company_id: int) -> bool:
    return product.get("companyId") == company_id or product.get("ownerCompanyId") == company_id


def sheet_visible_for_company(sheet: Dict[str, Any], company_id: int) -> bool:
    return (
        sheet.get("companyId") == company_id
        or sheet.get("ownerCompanyId") == company_id
        or company_id in (sheet.get("sharedCompanyIds") or [])
    )


def is_stock_tracked_product(product: Optional[Dict[str, Any]]) -> bool:
    return bool(product) and product.get("ignoreStock") is not True


def package_quantity(package: Dict[str, Any], control_unit: str) -> float:
    quantity = parse_decimal(package.get("packageQuantity")) or 0
    if quantity <= 0:
        return 0
    if control_unit in ("UNIT", "COMBO"):
        return quantity
    if control_unit == "MILLILITER":
        return quantity * 1000 if package.get("packageUnit") == "LITER" else quantity
    return quantity * 1000 if package.get("packageUnit") == "KILOGRAM" else quantity


def preferred_package(product: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    packages = [package for package in product.get("packages") or [] if package.get("isActive") is not False]
    if not packages:
        return None
    packages.sort(key=lambda package: (package.get("id") is None, package.get("id") or 0))
    return packages[0]


def ingredient_base_quantity(ingredient: Dict[str, Any]) -> float:
    return parse_decimal(ingredient.get("quantity")) or 0


def ingredient_yield_quantity(ingredient: Dict[str, Any]) -> float:
    return parse_decimal(ingredient.get("yieldQuantity")) or 0


def execution_final_yield(sheet: Dict[str, Any], products_by_id: Dict[str, Dict[str, Any]]) -> float:
    mixture_yield = 0.0
    for ingredient in sheet.get("ingredients") or []:
        if ingredient.get("isActive") is not True:
            continue
        product = products_by_id.get(ingredient.get("productId") or "")
        if product and product.get("excludeFromExecutionYield") is True:
            continue
        mixture_yield += ingredient_yield_quantity(ingredient)
    dilution = max(parse_decimal(sheet.get("dilutionRatePercentage")) or 0, 0)
    return mixture_yield + mixture_yield * (dilution / 100)


def effective_yield(sheet: Dict[str, Any], products_by_id: Dict[str, Dict[str, Any]], with_products: bool) -> float:
    kind = sheet.get("kind")
    if kind == "EXECUCAO":
        calculated = execution_final_yield(sheet, products_by_id if with_products else {})
        if calculated > 0:
            return calculated
        saved = parse_decimal(sheet.get("outputQuantity")) or 0
        if saved > 0:
            return saved
    if kind == "PREPARO":
        saved = parse_decimal(sheet.get("outputQuantity")) or 0
        if saved > 0:
            return saved
    include_garnishes = kind != "EXECUCAO"
    source = list(sheet.get("ingredients") or [])
    if include_garnishes:
        source += list(sheet.get("garnishIngredients") or [])
    ingredient_yield = sum(
        ingredient_yield_quantity(ingredient)
        for ingredient in source
        if ingredient.get("isActive") is True
    )
    if ingredient_yield > 0:
        return ingredient_yield
    return parse_decimal(sheet.get("outputQuantity")) or 0


def technical_sheet_base_yield(sheet: Dict[str, Any], products_by_id: Dict[str, Dict[str, Any]], with_products: bool) -> float:
    value = effective_yield(sheet, products_by_id, with_products)
    return value if value > 0 else 1


def stock_center_base_quantity(sheet: Dict[str, Any], products_by_id: Dict[str, Dict[str, Any]]) -> float:
    if sheet.get("outputUnit") == "UNIT":
        return 1
    portion_base = parse_decimal(sheet.get("portionSize")) or 0
    return portion_base if portion_base > 0 else technical_sheet_base_yield(sheet, products_by_id, True)


def build_consumptions_for_sale(
    sheet: Dict[str, Any],
    sold_quantity: float,
    products_by_id: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    import_base = technical_sheet_base_yield(sheet, products_by_id, False)
    panel_base = technical_sheet_base_yield(sheet, products_by_id, True)
    desired_yield = import_base * sold_quantity
    multiplier = desired_yield / panel_base if panel_base > 0 else sold_quantity
    rows = []
    for section_name, ingredients in (
        ("ingredients", sheet.get("ingredients") or []),
        ("garnishIngredients", sheet.get("garnishIngredients") or []),
    ):
        for ingredient in ingredients:
            if ingredient.get("isActive") is not True:
                continue
            product_id = str(ingredient.get("productId") or "").strip()
            if not product_id:
                continue
            product = products_by_id.get(product_id)
            rows.append(
                {
                    "section": section_name,
                    "source_sheet_id": sheet["id"],
                    "source_sheet_name": sheet["name"],
                    "source_sheet_kind": sheet["kind"],
                    "ingredient_product_id": product_id,
                    "ingredient_name": ingredient.get("productLabel") or (product or {}).get("name") or product_id,
                    "quantity": ingredient_base_quantity(ingredient) * multiplier,
                    "unit": unit_label((product or {}).get("controlUnit", "")),
                }
            )
    return rows


def build_inventory_key(kind: str, technical_sheet_id: Optional[int], product_id: str, service_item_id: str = "") -> str:
    if kind == "PREPARO":
        return f"PREPARO:{technical_sheet_id or ''}"
    if kind == "PRODUTO":
        return f"PRODUTO:{product_id}"
    if kind == "ITEM":
        return f"ITEM:{service_item_id}"
    return f"VENDA:{technical_sheet_id or ''}"


def is_operational_location(location: str) -> bool:
    normalized = normalize_text(location)
    return normalized in OPERATIONAL_LOCATIONS or normalized.startswith("SAIDA POR DESPERDICIO")


def is_outbound_location(location: str) -> bool:
    normalized = normalize_text(location)
    return normalized in OUTBOUND_LOCATIONS or normalized.startswith("SAIDA POR DESPERDICIO")


def current_inventory_balances(
    company_id: int,
    centers: List[Dict[str, Any]],
    inventories: List[Dict[str, Any]],
    sessions: List[Dict[str, Any]],
    counts: List[Dict[str, Any]],
) -> Tuple[Dict[int, str], Dict[Tuple[int, str], float]]:
    latest_date_by_center: Dict[int, str] = {}
    quantities: Dict[Tuple[int, str], float] = {}
    session_closed_at = {session["id"]: session.get("closedAt") or "" for session in sessions}
    relevant_center_ids = {
        center["id"] for center in centers if center.get("companyId") == company_id and center.get("isActive") is not False
    }
    relevant_center_ids.update(record["stockCenterId"] for record in counts if record.get("companyId") == company_id)
    relevant_center_ids.update(record["stockCenterId"] for record in inventories if record.get("companyId") == company_id)

    for center_id in sorted(relevant_center_ids):
        center_counts = [
            record
            for record in counts
            if record.get("companyId") == company_id and record.get("stockCenterId") == center_id
        ]
        center_inventories = [
            record
            for record in inventories
            if record.get("companyId") == company_id
            and record.get("stockCenterId") == center_id
            and record.get("isClosed") is True
        ]
        events = []
        for inventory in center_inventories:
            events.append(
                {
                    "kind": "INVENTORY",
                    "countedAt": inventory.get("countedAt") or "",
                    "eventTimestamp": inventory.get("startedAt") or inventory.get("closedAt") or inventory.get("countedAt") or "",
                    "sortId": inventory.get("id") or 0,
                    "inventory": inventory,
                }
            )
        for record in center_counts:
            if is_operational_location(record.get("storageLocation") or ""):
                events.append(
                    {
                        "kind": "RECORD",
                        "countedAt": record.get("countedAt") or "",
                        "eventTimestamp": session_closed_at.get(record.get("sessionId"), "") or record.get("countedAt") or "",
                        "sortId": record.get("id") or 0,
                        "record": record,
                    }
                )
        events.sort(key=lambda event: (event["countedAt"], event["eventTimestamp"], event["sortId"]))

        balance_by_key: Dict[str, float] = {}
        for event in events:
            if event["kind"] == "INVENTORY":
                inventory_id = event["inventory"].get("id")
                counted_quantities: Dict[str, float] = defaultdict(float)
                for record in center_counts:
                    if record.get("inventoryId") != inventory_id:
                        continue
                    if is_operational_location(record.get("storageLocation") or ""):
                        continue
                    kind = "PREPARO" if record.get("technicalSheetKind") == "VENDA" else record.get("technicalSheetKind")
                    key = build_inventory_key(kind, record.get("technicalSheetId"), record.get("productId") or "", record.get("serviceItemId") or "")
                    counted_quantities[key] += parse_decimal(record.get("totalCountedQuantity")) or 0
                all_keys = set(balance_by_key) | set(counted_quantities)
                balance_by_key = {key: counted_quantities.get(key, 0) for key in all_keys}
                latest_date_by_center[center_id] = event["countedAt"]
                continue

            record = event["record"]
            kind = "PREPARO" if record.get("technicalSheetKind") == "VENDA" else record.get("technicalSheetKind")
            key = build_inventory_key(kind, record.get("technicalSheetId"), record.get("productId") or "", record.get("serviceItemId") or "")
            movement_quantity = abs(parse_decimal(record.get("totalCountedQuantity")) or 0)
            current = balance_by_key.get(key, 0)
            balance_by_key[key] = current - movement_quantity if is_outbound_location(record.get("storageLocation") or "") else current + movement_quantity
            latest_date_by_center[center_id] = event["countedAt"]

        for key, quantity in balance_by_key.items():
            quantities[(center_id, key)] = quantity

    return latest_date_by_center, quantities


def resolve_drink_sheet(
    drink_name: str,
    sheets: List[Dict[str, Any]],
    company_id: int,
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    normalized = normalize_text(drink_name)
    candidates = [
        sheet
        for sheet in sheets
        if sheet.get("isActive") is True
        and sheet.get("kind") in ("EXECUCAO", "VENDA")
        and sheet_visible_for_company(sheet, company_id)
        and normalize_text(sheet.get("name")) == normalized
    ]
    if len(candidates) == 1:
        return candidates[0], []
    if len(candidates) > 1:
        return None, [f"Drink '{drink_name}' tem multiplas fichas comerciais/executivas: {[c['id'] for c in candidates]}"]

    loose = [
        sheet
        for sheet in sheets
        if sheet.get("isActive") is True
        and sheet.get("kind") in ("EXECUCAO", "VENDA")
        and sheet_visible_for_company(sheet, company_id)
        and (normalize_text(sheet.get("name")) in normalized or normalized in normalize_text(sheet.get("name")))
    ]
    if len(loose) == 1:
        return loose[0], [f"Drink '{drink_name}' resolvido por aproximacao para ficha {loose[0]['id']} - {loose[0]['name']}."]
    if len(loose) > 1:
        candidates = [f"{candidate['id']}:{candidate['name']}" for candidate in loose[:8]]
        return None, [f"Drink '{drink_name}' sem match unico; candidatos aproximados: {candidates}"]
    return None, [f"Drink '{drink_name}' nao possui ficha EXECUCAO/VENDA ativa encontrada."]


def resolve_product(
    product_name: str,
    products: List[Dict[str, Any]],
    company_id: int,
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    normalized = normalize_text(product_name)
    target_name = CLOSED_BEVERAGE_ALIASES.get(normalized, normalized)
    target = normalize_text(target_name)
    candidates = [
        product
        for product in products
        if product.get("isActive") is True
        and product_visible_for_company(product, company_id)
        and normalize_text(product.get("name")) == target
    ]
    if len(candidates) == 1:
        notes = []
        if target != normalized:
            notes.append(f"Produto '{product_name}' resolvido por alias para '{candidates[0]['name']}'.")
        return candidates[0], notes
    if len(candidates) > 1:
        return None, [f"Produto '{product_name}' tem multiplos cadastros exatos: {[c['id'] for c in candidates]}"]

    loose = [
        product
        for product in products
        if product.get("isActive") is True
        and product_visible_for_company(product, company_id)
        and (target in normalize_text(product.get("name")) or normalize_text(product.get("name")) in target)
    ]
    if len(loose) == 1:
        return loose[0], [f"Produto '{product_name}' resolvido por aproximacao para {loose[0]['id']} - {loose[0]['name']}."]
    if len(loose) > 1:
        candidates = [f"{candidate['id']}:{candidate['name']}" for candidate in loose[:8]]
        return None, [f"Produto '{product_name}' sem match unico; candidatos: {candidates}"]
    return None, [f"Produto '{product_name}' nao encontrado no cadastro ativo."]


def resolve_sale_sheet_for_closed_product(
    projection_name: str,
    product: Dict[str, Any],
    sheets: List[Dict[str, Any]],
    company_id: int,
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    product_id = product.get("id") or ""
    linked_candidates = [
        sheet
        for sheet in sheets
        if sheet.get("kind") == "VENDA"
        and sheet.get("isActive") is True
        and sheet_visible_for_company(sheet, company_id)
        and any(
            ingredient.get("isActive") is True and ingredient.get("productId") == product_id
            for ingredient in (sheet.get("ingredients") or []) + (sheet.get("garnishIngredients") or [])
        )
    ]
    if len(linked_candidates) == 1:
        return linked_candidates[0], []
    if len(linked_candidates) > 1:
        candidates = [f"{candidate['id']}:{candidate['name']}" for candidate in linked_candidates[:8]]
        return None, [f"Bebida fechada '{projection_name}' tem multiplas fichas VENDA consumindo {product['name']}: {candidates}"]

    normalized_projection = normalize_text(projection_name)
    alias_name = CLOSED_BEVERAGE_ALIASES.get(normalized_projection, normalized_projection)
    name_candidates = [
        sheet
        for sheet in sheets
        if sheet.get("kind") == "VENDA"
        and sheet.get("isActive") is True
        and sheet_visible_for_company(sheet, company_id)
        and (
            normalize_text(sheet.get("name")) == alias_name
            or normalize_text(sheet.get("name")) == normalized_projection
        )
    ]
    if len(name_candidates) == 1:
        return name_candidates[0], [f"Bebida fechada '{projection_name}' resolvida por nome para ficha VENDA {name_candidates[0]['id']}."]
    if len(name_candidates) > 1:
        candidates = [f"{candidate['id']}:{candidate['name']}" for candidate in name_candidates[:8]]
        return None, [f"Bebida fechada '{projection_name}' tem multiplas fichas VENDA por nome: {candidates}"]

    package = preferred_package(product)
    package_base_quantity = package_quantity(package, product.get("controlUnit") or "") if package else 0
    return None, [
        "Bebida fechada "
        f"'{projection_name}' nao tem ficha VENDA. Criar ficha VENDA resumida consumindo 1 embalagem de "
        f"{product['name']} ({format_decimal(package_base_quantity)} {unit_label(product.get('controlUnit') or '')}) antes de importar vendas."
    ]


def resolve_preparation_sheet_by_product_id(
    product_id: str,
    sheets: List[Dict[str, Any]],
    company_id: int,
) -> Optional[Dict[str, Any]]:
    candidates = [
        sheet
        for sheet in sheets
        if sheet.get("kind") == "PREPARO"
        and sheet.get("isActive") is True
        and sheet.get("productId") == product_id
        and sheet_visible_for_company(sheet, company_id)
    ]
    candidates.sort(key=lambda sheet: (0 if sheet.get("ownerCompanyId") == company_id else 1, sheet.get("id") or 0))
    return candidates[0] if candidates else None


def aggregate_required_targets(
    consumptions: List[Dict[str, Any]],
    closed_targets: List[Dict[str, Any]],
    sheets: List[Dict[str, Any]],
    products_by_id: Dict[str, Dict[str, Any]],
    company_id: int,
) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    targets: Dict[str, Dict[str, Any]] = {}
    inconsistencies = []

    def ensure_target(key: str, data: Dict[str, Any]) -> Dict[str, Any]:
        if key not in targets:
            targets[key] = data
        return targets[key]

    for consumption in consumptions:
        product_id = consumption["ingredient_product_id"]
        product = products_by_id.get(product_id)
        if not product:
            inconsistencies.append(
                f"Ingrediente '{consumption['ingredient_name']}' da ficha {consumption['source_sheet_id']} aponta para produto inexistente: {product_id}."
            )
            continue
        if not is_stock_tracked_product(product):
            continue
        prep_sheet = resolve_preparation_sheet_by_product_id(product_id, sheets, company_id)
        if prep_sheet:
            key = build_inventory_key("PREPARO", prep_sheet["id"], prep_sheet.get("productId") or "")
            target = ensure_target(
                key,
                {
                    "key": key,
                    "kind": "PREPARO",
                    "technical_sheet_id": prep_sheet["id"],
                    "product_id": prep_sheet.get("productId") or product_id,
                    "name": prep_sheet["name"],
                    "family": prep_sheet.get("family") or "",
                    "unit": unit_label(prep_sheet.get("outputUnit") or ""),
                    "required_quantity": 0.0,
                    "sources": [],
                    "is_closed_beverage": False,
                },
            )
        else:
            key = build_inventory_key("PRODUTO", None, product_id)
            target = ensure_target(
                key,
                {
                    "key": key,
                    "kind": "PRODUTO",
                    "technical_sheet_id": None,
                    "product_id": product_id,
                    "name": product.get("name") or product_id,
                    "family": product.get("family") or "",
                    "unit": unit_label(product.get("controlUnit") or ""),
                    "required_quantity": 0.0,
                    "sources": [],
                    "is_closed_beverage": False,
                },
            )
        quantity = consumption["quantity"]
        target["required_quantity"] += quantity
        target["sources"].append(
            {
                "source": consumption["source_sheet_name"],
                "sheet_id": consumption["source_sheet_id"],
                "quantity": quantity,
            }
        )

    for closed in closed_targets:
        product = closed["product"]
        package = preferred_package(product)
        if not package:
            inconsistencies.append(f"Bebida fechada '{closed['projection_name']}' sem embalagem ativa para converter unidade de compra em unidade de estoque.")
            continue
        package_base_quantity = package_quantity(package, product.get("controlUnit") or "")
        if package_base_quantity <= 0:
            inconsistencies.append(f"Bebida fechada '{closed['projection_name']}' possui embalagem sem quantidade normalizada.")
            continue
        key = build_inventory_key("PRODUTO", None, product["id"])
        target = ensure_target(
            key,
            {
                "key": key,
                "kind": "PRODUTO",
                "technical_sheet_id": None,
                "product_id": product["id"],
                "name": product.get("name") or product["id"],
                "family": product.get("family") or "",
                "unit": unit_label(product.get("controlUnit") or ""),
                "required_quantity": 0.0,
                "sources": [],
                "is_closed_beverage": True,
            },
        )
        required = closed["target_units"] * package_base_quantity
        target["required_quantity"] += required
        target["sources"].append(
            {
                "source": "BEBIDA FECHADA",
                "projection_name": closed["projection_name"],
                "target_units": closed["target_units"],
                "package_quantity": package_base_quantity,
                "quantity": required,
            }
        )

    return targets, inconsistencies


def requested_base_quantity(line: Dict[str, Any], products_by_id: Dict[str, Dict[str, Any]], sheets_by_id: Dict[int, Dict[str, Any]]) -> float:
    requested = parse_decimal(line.get("requestedQuantity")) or 0
    if requested <= 0:
        return 0
    kind = line.get("kind")
    if kind == "PREPARO" and isinstance(line.get("technicalSheetId"), int):
        sheet = sheets_by_id.get(line["technicalSheetId"])
        if not sheet:
            return requested
        return requested * stock_center_base_quantity(sheet, products_by_id)
    if kind == "PRODUTO" and line.get("packageId") is not None:
        product = products_by_id.get(line.get("productId") or "")
        if not product:
            return requested
        package = next((pkg for pkg in product.get("packages") or [] if pkg.get("id") == line.get("packageId")), None)
        return requested * package_quantity(package, product.get("controlUnit") or "") if package else requested
    return requested


def active_commitments_by_supplier(
    requisitions: List[Dict[str, Any]],
    products_by_id: Dict[str, Dict[str, Any]],
    sheets_by_id: Dict[int, Dict[str, Any]],
) -> Dict[Tuple[int, str], float]:
    commitments: Dict[Tuple[int, str], float] = defaultdict(float)
    for requisition in requisitions:
        if requisition.get("status") in ("CANCELLED", "RECEIVED"):
            continue
        supplier_id = requisition.get("supplyCenterId")
        if not isinstance(supplier_id, int):
            continue
        for line in requisition.get("lines") or []:
            kind = line.get("kind")
            key = build_inventory_key(
                "PREPARO" if kind == "VENDA" else kind,
                line.get("technicalSheetId"),
                line.get("productId") or "",
                line.get("serviceItemId") or "",
            )
            commitments[(supplier_id, key)] += requested_base_quantity(line, products_by_id, sheets_by_id)
    return commitments


def build_recommendations(
    targets: Dict[str, Dict[str, Any]],
    centers: List[Dict[str, Any]],
    bar_center_id: int,
    balances: Dict[Tuple[int, str], float],
    commitments: Dict[Tuple[int, str], float],
) -> List[Dict[str, Any]]:
    centers_by_id = {center["id"]: center for center in centers}
    recommendations = []
    for key, target in sorted(targets.items(), key=lambda item: item[1]["name"]):
        required = target["required_quantity"]
        bar_stock = balances.get((bar_center_id, key), 0.0)
        need = max(required - bar_stock, 0.0)

        transfer_lines = []
        remaining = need
        other_available_total = 0.0
        for center in sorted(centers, key=lambda item: item.get("name") or ""):
            center_id = center["id"]
            if center_id == bar_center_id or center.get("isActive") is False:
                continue
            stock = balances.get((center_id, key), 0.0)
            committed = commitments.get((center_id, key), 0.0)
            available = max(stock - committed, 0.0)
            if available <= 0:
                continue
            other_available_total += available
            transfer = min(remaining, available)
            if transfer > 0:
                transfer_lines.append(
                    {
                        "center_id": center_id,
                        "center_name": centers_by_id.get(center_id, {}).get("name", f"CENTRO {center_id}"),
                        "current_stock": stock,
                        "committed_active_requisitions": committed,
                        "available": available,
                        "suggested_transfer": transfer,
                    }
                )
                remaining -= transfer

        purchase = max(remaining, 0.0)
        projected_balance = bar_stock - required
        recommendations.append(
            {
                **target,
                "required_quantity": required,
                "bar_stock": bar_stock,
                "projected_bar_balance_after_consumption": projected_balance,
                "gross_need_for_projection": need,
                "available_other_centers": other_available_total,
                "suggested_internal_requisition": need - purchase,
                "purchase_required": purchase,
                "transfer_lines": transfer_lines,
            }
        )
    return recommendations


def build_report(args: argparse.Namespace) -> Dict[str, Any]:
    projection = read_projection(args.workbook)
    company_id = args.company_id
    if args.sales_start_date or args.sales_end_date:
        if not args.sales_start_date or not args.sales_end_date:
            raise RuntimeError("Informe sales-start-date e sales-end-date juntos.")
        dated_projection = fixed_sales_dates(args.sales_start_date, args.sales_end_date)
    else:
        dated_projection = projection_dates(projection["crowd"], args.first_projection_date)
    simulation_id = f"madre-demand-projection-{dt.datetime.now(dt.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"

    companies = api_get(args.api_base, "/companies").get("companies", [])
    centers = api_get(args.api_base, "/stock-centers", {"companyId": company_id}).get("stockCenters", [])
    products = api_get(args.api_base, "/products", {"companyId": company_id}).get("products", [])
    sheets = api_get(args.api_base, "/technical-sheets", {"companyId": company_id}).get("technicalSheets", [])
    inventories = api_get(args.api_base, "/inventories", {"companyId": company_id}).get("inventoryRecords", [])
    sessions = api_get(args.api_base, "/inventory-count-sessions", {"companyId": company_id}).get("inventoryCountSessions", [])
    counts = api_get(args.api_base, "/inventory-counts", {"companyId": company_id}).get("inventoryCounts", [])
    requisitions = api_get(args.api_base, "/requisitions", {"companyId": company_id}).get("requisitions", [])
    sales_batches = api_get(args.api_base, "/sales-import-batches", {"companyId": company_id}).get("batches", [])
    sales_consumptions = api_get(args.api_base, "/sales-consumptions", {"companyId": company_id}).get("consumptions", [])

    company = next((item for item in companies if item.get("id") == company_id), None)
    bar_center = next((center for center in centers if normalize_text(center.get("name")) == normalize_text(args.center_name)), None)
    if not bar_center:
        raise RuntimeError(f"Centro de estoque nao encontrado: {args.center_name}")

    products_by_id = {product["id"]: product for product in products}
    sheets_by_id = {sheet["id"]: sheet for sheet in sheets}
    inconsistencies: List[str] = []
    resolution_notes: List[str] = []

    simulated_sales = []
    consumptions = []
    for item in projection["authorials"]:
        sheet, notes = resolve_drink_sheet(item["drink"], sheets, company_id)
        resolution_notes.extend(notes)
        if not sheet:
            inconsistencies.extend(notes)
            continue
        projected_quantity = item["capacity_total"]
        fake_quantity = projected_quantity * args.fake_demand_factor
        simulated_sales.append(
            {
                "group": "AUTORAL",
                "drink": item["drink"],
                "sheet_id": sheet["id"],
                "product_id": sheet.get("productId") or "",
                "projected_quantity": projected_quantity,
                "quantity": fake_quantity,
                "promotional_quantity_included": 0,
                "priority": item["priority"],
            }
        )
        consumptions.extend(build_consumptions_for_sale(sheet, fake_quantity, products_by_id))

    authorial_total = sum(row["quantity"] for row in simulated_sales if row["group"] == "AUTORAL")
    if authorial_total > 0:
        remaining_promo = PROMOTIONAL_AUTHORIAL_UNITS
        for row in sorted(
            [row for row in simulated_sales if row["group"] == "AUTORAL"],
            key=lambda item: (0 if normalize_text(item.get("priority")) == "ALTA" else 1, item["drink"]),
        ):
            assigned = min(row["quantity"], remaining_promo)
            row["promotional_quantity_included"] = assigned
            remaining_promo -= assigned
            if remaining_promo <= 0:
                break

    for item in projection["classics"]:
        sheet, notes = resolve_drink_sheet(item["drink"], sheets, company_id)
        resolution_notes.extend(notes)
        if not sheet:
            inconsistencies.extend(notes)
            continue
        projected_quantity = item["capacity"]
        fake_quantity = projected_quantity * args.fake_demand_factor
        simulated_sales.append(
            {
                "group": "CLASSICO",
                "drink": item["drink"],
                "sheet_id": sheet["id"],
                "product_id": sheet.get("productId") or "",
                "projected_quantity": projected_quantity,
                "quantity": fake_quantity,
                "promotional_quantity_included": 0,
                "priority": "",
            }
        )
        consumptions.extend(build_consumptions_for_sale(sheet, fake_quantity, products_by_id))

    direct_closed_targets = []
    closed_beverage_sale_sheet_requirements = []
    for item in projection["closed_beverages"]:
        product, notes = resolve_product(item["product"], products, company_id)
        resolution_notes.extend(notes)
        if not product:
            inconsistencies.extend(notes)
            continue
        sale_sheet, sale_notes = resolve_sale_sheet_for_closed_product(item["product"], product, sheets, company_id)
        resolution_notes.extend(sale_notes)
        if sale_sheet:
            projected_quantity = item["target_units"]
            fake_quantity = projected_quantity * args.fake_demand_factor
            simulated_sales.append(
                {
                    "group": "BEBIDA FECHADA",
                    "drink": item["product"],
                    "sheet_id": sale_sheet["id"],
                    "product_id": sale_sheet.get("productId") or "",
                    "projected_quantity": projected_quantity,
                    "quantity": fake_quantity,
                    "promotional_quantity_included": 0,
                    "priority": "",
                }
            )
            consumptions.extend(build_consumptions_for_sale(sale_sheet, fake_quantity, products_by_id))
        else:
            inconsistencies.extend(sale_notes)
            package = preferred_package(product)
            package_base_quantity = package_quantity(package, product.get("controlUnit") or "") if package else 0
            closed_beverage_sale_sheet_requirements.append(
                {
                    "projection_name": item["product"],
                    "product_id": product["id"],
                    "product_name": product["name"],
                    "suggested_sale_sheet_name": item["product"].upper(),
                    "required_ingredient_quantity_per_unit": package_base_quantity,
                    "unit": unit_label(product.get("controlUnit") or ""),
                }
            )
            direct_closed_targets.append(
                {
                    "projection_name": item["product"],
                    "projected_target_units": item["target_units"],
                    "target_units": item["target_units"] * args.fake_demand_factor,
                    "product": product,
                }
            )

    required_targets, target_inconsistencies = aggregate_required_targets(
        consumptions,
        direct_closed_targets,
        sheets,
        products_by_id,
        company_id,
    )
    inconsistencies.extend(target_inconsistencies)

    latest_dates, balances = current_inventory_balances(company_id, centers, inventories, sessions, counts)
    commitments = active_commitments_by_supplier(requisitions, products_by_id, sheets_by_id)
    recommendations = build_recommendations(required_targets, centers, bar_center["id"], balances, commitments)
    fake_import_rows = []
    for sale in simulated_sales:
        identifier = sale.get("product_id") or str(sale["sheet_id"])
        for date_key, quantity in distribute_quantity(sale["quantity"], dated_projection):
            fake_import_rows.append(
                {
                    "DATA": date_key,
                    "IDENTIFICADOR": identifier,
                    "QUANTIDADE": quantity,
                    "ITEM": sale["drink"],
                    "GRUPO": sale["group"],
                    "FICHA_ID": sale["sheet_id"],
                    "SIMULATION_ID": simulation_id,
                }
            )

    report = {
        "simulation_id": simulation_id,
        "mode": "DRY_RUN_READ_ONLY",
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "api_base": args.api_base,
        "workbook": str(Path(args.workbook).resolve()),
        "strategy": {
            "authorials": "Use CAPACIDADE TOTAL as operational target; the 140 promotional opening drinks are tagged as included demand, not added on top.",
            "classics": "Use planned capacity from workbook, normally 15 units per classic.",
            "closed_beverages": "Use ESTOQUE RECOMENDADO as direct target in packages/units, converted to product control unit.",
            "purchase_logic": "Need for Bar = max(projected requirement - current Bar stock, 0); internal supply consumes available stock in other centers before purchase residual.",
            "safety_margin_percent_reference": SAFETY_MARGIN_PERCENT,
            "fake_demand_factor": args.fake_demand_factor,
            "fake_demand_factor_reason": "Projection already includes operational buffer; imported fake sales are reduced before the webapp applies its own safety margin.",
            "fake_sales_date_range": {
                "start": dated_projection[0]["date"] if dated_projection else "",
                "end": dated_projection[-1]["date"] if dated_projection else "",
            },
        },
        "company": {
            "id": company_id,
            "name": (company or {}).get("tradeName")
            or (company or {}).get("name")
            or (company or {}).get("legalName")
            or "CASA DE MI MADRE",
        },
        "target_center": {
            "id": bar_center["id"],
            "name": bar_center["name"],
            "latest_inventory_date": latest_dates.get(bar_center["id"], ""),
        },
        "source_state": {
            "stock_centers": [{"id": c["id"], "name": c["name"], "isDistributor": c.get("isDistributor"), "suppliedCenterIds": c.get("suppliedCenterIds") or []} for c in centers],
            "closed_inventories": len([inventory for inventory in inventories if inventory.get("isClosed") is True]),
            "inventory_counts": len(counts),
            "sales_import_batches": len(sales_batches),
            "sales_consumptions": len(sales_consumptions),
            "requisitions_total": len(requisitions),
            "active_requisitions": len([req for req in requisitions if req.get("status") not in ("CANCELLED", "RECEIVED")]),
        },
        "demand_summary": {
            "crowd_min": sum(row["minimum"] for row in projection["crowd"]),
            "crowd_max": sum(row["maximum"] for row in projection["crowd"]),
            "authorial_units": sum(row["quantity"] for row in simulated_sales if row["group"] == "AUTORAL"),
            "authorial_projected_units_before_factor": sum(row["projected_quantity"] for row in simulated_sales if row["group"] == "AUTORAL"),
            "classic_units": sum(row["quantity"] for row in simulated_sales if row["group"] == "CLASSICO"),
            "classic_projected_units_before_factor": sum(row["projected_quantity"] for row in simulated_sales if row["group"] == "CLASSICO"),
            "closed_beverage_units": sum(row["target_units"] for row in direct_closed_targets)
            + sum(row["quantity"] for row in simulated_sales if row["group"] == "BEBIDA FECHADA"),
            "closed_beverage_projected_units_before_factor": sum(row["projected_target_units"] for row in direct_closed_targets)
            + sum(row["projected_quantity"] for row in simulated_sales if row["group"] == "BEBIDA FECHADA"),
            "promotional_authorial_units_included": sum(row["promotional_quantity_included"] for row in simulated_sales),
        },
        "dated_projection": dated_projection,
        "simulated_sales": simulated_sales,
        "fake_sales_import_rows": fake_import_rows,
        "closed_beverage_direct_targets_without_sale_sheet": [
            {
                "projection_name": item["projection_name"],
                "product_id": item["product"]["id"],
                "product_name": item["product"]["name"],
                "target_units": item["target_units"],
                "projected_target_units": item["projected_target_units"],
            }
            for item in direct_closed_targets
        ],
        "closed_beverage_sale_sheet_requirements": closed_beverage_sale_sheet_requirements,
        "ingredient_consumption": [
            {
                "product_id": product_id,
                "product_name": (products_by_id.get(product_id) or {}).get("name") or product_id,
                "quantity": quantity,
                "unit": unit_label((products_by_id.get(product_id) or {}).get("controlUnit", "")),
            }
            for product_id, quantity in sorted(
                {
                    product_id: sum(row["quantity"] for row in consumptions if row["ingredient_product_id"] == product_id)
                    for product_id in {row["ingredient_product_id"] for row in consumptions}
                }.items(),
                key=lambda item: (products_by_id.get(item[0]) or {}).get("name") or item[0],
            )
        ],
        "recommendations": recommendations,
        "inconsistencies": sorted(set(inconsistencies)),
        "resolution_notes": sorted(set(note for note in resolution_notes if note not in inconsistencies)),
    }
    return report


def write_json(report: Dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{report['simulation_id']}.json"
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_xlsx(report: Dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{report['simulation_id']}.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Resumo"
    ws.append(["simulation_id", report["simulation_id"]])
    ws.append(["modo", report["mode"]])
    ws.append(["empresa", f"{report['company']['id']} - {report['company']['name']}"])
    ws.append(["centro", f"{report['target_center']['id']} - {report['target_center']['name']}"])
    ws.append(["publico_min", report["demand_summary"]["crowd_min"]])
    ws.append(["publico_max", report["demand_summary"]["crowd_max"]])
    ws.append(["autorais_fake", report["demand_summary"]["authorial_units"]])
    ws.append(["autorais_projecao", report["demand_summary"]["authorial_projected_units_before_factor"]])
    ws.append(["classicos_fake", report["demand_summary"]["classic_units"]])
    ws.append(["classicos_projecao", report["demand_summary"]["classic_projected_units_before_factor"]])
    ws.append(["bebidas_fechadas_fake_un", report["demand_summary"]["closed_beverage_units"]])
    ws.append(["bebidas_fechadas_projecao_un", report["demand_summary"]["closed_beverage_projected_units_before_factor"]])
    ws.append(["fator_fake", report["strategy"]["fake_demand_factor"]])
    ws.append(["inconsistencias", len(report["inconsistencies"])])

    ws = workbook.create_sheet("Vendas simuladas")
    ws.append(["grupo", "drink", "sheet_id", "product_id", "quantidade_fake", "quantidade_projecao", "promo_incluida", "prioridade"])
    for row in report["simulated_sales"]:
        ws.append([row["group"], row["drink"], row["sheet_id"], row["product_id"], row["quantity"], row["projected_quantity"], row["promotional_quantity_included"], row["priority"]])

    ws = workbook.create_sheet("Importacao vendas")
    ws.append(["DATA", "IDENTIFICADOR", "QUANTIDADE", "ITEM", "GRUPO", "FICHA_ID", "SIMULATION_ID"])
    for row in report["fake_sales_import_rows"]:
        ws.append([
            row["DATA"],
            row["IDENTIFICADOR"],
            row["QUANTIDADE"],
            row["ITEM"],
            row["GRUPO"],
            row["FICHA_ID"],
            row["SIMULATION_ID"],
        ])

    ws = workbook.create_sheet("Consumo insumos")
    ws.append(["product_id", "produto", "quantidade", "unidade"])
    for row in report["ingredient_consumption"]:
        ws.append([row["product_id"], row["product_name"], row["quantity"], row["unit"]])

    ws = workbook.create_sheet("Requisicao compra")
    ws.append([
        "tipo",
        "item",
        "product_id",
        "technical_sheet_id",
        "necessidade_bar",
        "estoque_bar",
        "saldo_proj_bar",
        "disp_outros",
        "req_interna",
        "compra",
        "unidade",
        "fontes_transferencia",
    ])
    for row in report["recommendations"]:
        ws.append([
            row["kind"],
            row["name"],
            row["product_id"],
            row["technical_sheet_id"],
            row["gross_need_for_projection"],
            row["bar_stock"],
            row["projected_bar_balance_after_consumption"],
            row["available_other_centers"],
            row["suggested_internal_requisition"],
            row["purchase_required"],
            row["unit"],
            "; ".join(
                f"{line['center_name']}={format_decimal(line['suggested_transfer'])}"
                for line in row["transfer_lines"]
            ),
        ])

    ws = workbook.create_sheet("Inconsistencias")
    ws.append(["tipo", "mensagem"])
    for message in report["inconsistencies"]:
        ws.append(["ERRO", message])
    for message in report["resolution_notes"]:
        ws.append(["NOTA", message])

    ws = workbook.create_sheet("Fichas venda pendentes")
    ws.append(["projecao", "produto_id", "produto", "nome_ficha_sugerido", "quantidade_por_venda", "unidade"])
    for row in report["closed_beverage_sale_sheet_requirements"]:
        ws.append([
            row["projection_name"],
            row["product_id"],
            row["product_name"],
            row["suggested_sale_sheet_name"],
            row["required_ingredient_quantity_per_unit"],
            row["unit"],
        ])

    workbook.save(path)
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Dry-run Casa de mi Madre demand projection.")
    parser.add_argument("--api-base", default=DEFAULT_API_BASE)
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--company-id", type=int, default=DEFAULT_COMPANY_ID)
    parser.add_argument("--center-name", default=DEFAULT_CENTER_NAME)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--first-projection-date", default="")
    parser.add_argument("--sales-start-date", default="")
    parser.add_argument("--sales-end-date", default="")
    parser.add_argument("--fake-demand-factor", type=float, default=DEFAULT_FAKE_DEMAND_FACTOR)
    args = parser.parse_args()

    report = build_report(args)
    output_dir = Path(args.output_dir)
    json_path = write_json(report, output_dir)
    xlsx_path = write_xlsx(report, output_dir)

    top_purchase = sorted(report["recommendations"], key=lambda row: row["purchase_required"], reverse=True)[:10]
    print(f"simulation_id={report['simulation_id']}")
    print(f"json={json_path}")
    print(f"xlsx={xlsx_path}")
    print(
        "summary "
        f"publico={format_decimal(report['demand_summary']['crowd_min'])}-{format_decimal(report['demand_summary']['crowd_max'])} "
        f"autorais_fake={format_decimal(report['demand_summary']['authorial_units'])} "
        f"classicos_fake={format_decimal(report['demand_summary']['classic_units'])} "
        f"bebidas_fechadas_fake_un={format_decimal(report['demand_summary']['closed_beverage_units'])} "
        f"fator={format_decimal(report['strategy']['fake_demand_factor'])}"
    )
    print(f"inconsistencies={len(report['inconsistencies'])}")
    print("top_purchase_required:")
    for row in top_purchase:
        if row["purchase_required"] <= 0:
            continue
        print(f"- {row['name']}: compra={format_decimal(row['purchase_required'])} {row['unit']} req_interna={format_decimal(row['suggested_internal_requisition'])} {row['unit']}")


if __name__ == "__main__":
    main()
